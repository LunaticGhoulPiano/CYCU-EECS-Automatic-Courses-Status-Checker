# -*- coding: utf-8 -*-
import os
import re
import json
import openpyxl
import pandas as pd

class DuplicateProgramError(Exception):
    def __init__(self, message):
        super().__init__(message)
        self.message = message

def generate_table(path, enroll_year):
    # set basic types of graduate credits
    graduate_credits = {
        '基本知能': 6,
        '通識基礎必修': 14,
        '學系必修': 66,
        '學系選修': 14,
        '通識延伸選修':12,
        '自由選修': 14,
        '其他選修': 0
    }
    if enroll_year == '108':
        graduate_credits['學系選修'] = 16
        graduate_credits['通識延伸選修'] = 14
        graduate_credits['自由選修'] = 12
    elif enroll_year == '109':
        graduate_credits['通識延伸選修'] = 14
    json_dict = { '畢業應修最低學分數': graduate_credits }
    
    # set courses and properties of each basic type of graduate credit
    for key in graduate_credits:
        temp_dict = {}
        if key == '基本知能':
            temp_dict = {
                '英文(一)': {
                    '期程': '半',
                    '學分數': 1,
                    '修習時間': f'{enroll_year}1 / 大一上'
                },
                '英文(二)': {
                    '期程': '半',
                    '學分數': 1,
                    '修習時間': f'{enroll_year}2 / 大一下'
                },
                '實用英文(一)': {
                    '期程': '半',
                    '學分數': 1,
                    '修習時間': f'{int(enroll_year)+1}1 / 大二上' # not sure
                },
                '實用英文(二)': {
                    '期程': '半',
                    '學分數': 1,
                    '修習時間': f'{int(enroll_year)+1}2 / 大二下' # not sure
                },
                '英語聽講(一)': {
                    '期程': '半',
                    '學分數': 1,
                    '修習時間': f'{enroll_year}1 / 大一上'
                },
                '英語聽講(二)': {
                    '期程': '半',
                    '學分數': 1,
                    '修習時間': f'{enroll_year}2 / 大一下'
                },
                '體育一': {
                    '期程': '半',
                    '學分數': 1,
                    '修習時間': f'{enroll_year}1 / 大一上'
                },
                '體育興趣一': {
                    '期程': '半',
                    '學分數': 0,
                    '修習時間': None
                },
                '體育興趣二': {
                    '期程': '半',
                    '學分數': 0,
                    '修習時間': None
                },
                '體育興趣三': {
                    '期程': '半',
                    '學分數': 0,
                    '修習時間': None
                }
            }
        elif key == '通識基礎必修':
            temp_dict = {
                '天': {
                    '宗教哲學': {
                        '期程': '半',
                        '學分數': 2,
                        '修習時間': None,
                        '性質': '宗哲'
                    },
                    '人生哲學': {
                        '期程': '半',
                        '學分數': 2,
                        '修習時間': None,
                        '性質': '人哲'
                    }
                },
                '人': {
                    '台灣政治與民主': {
                        '期程': '半',
                        '學分數': 2,
                        '修習時間': None,
                        '性質': '公民'
                    },
                    '法律與現代生活': {
                        '期程': '半',
                        '學分數': 2,
                        '修習時間': None,
                        '性質': '公民'
                    },
                    '當代人權議題與挑戰': {
                        '期程': '半',
                        '學分數': 2,
                        '修習時間': None,
                        '性質': '公民'
                    },
                    '生活社會學': {
                        '期程': '半',
                        '學分數': 2,
                        '修習時間': None,
                        '性質': '公民'
                    },
                    '全球化大議題': {
                        '期程': '半',
                        '學分數': 2,
                        '修習時間': None,
                        '性質': '公民'
                    },
                    '經濟學的世界': {
                        '期程': '半',
                        '學分數': 2,
                        '修習時間': None,
                        '性質': '公民'
                    },
                    '區域文明史': {
                        '期程': '半',
                        '學分數': 2,
                        '修習時間': None,
                        '性質': '歷史'
                    },
                    '文化思想史': {
                        '期程': '半',
                        '學分數': 2,
                        '修習時間': None,
                        '性質': '歷史'
                    }
                },
                '物': {
                    '運算思維與程式設計': {
                        '期程': '半',
                        '學分數': 2,
                        '修習時間': f'{enroll_year}1 / 大一上',
                    },
                    '自然科學與人工智慧導論': {
                        '期程': '半',
                        '學分數': 2,
                        '修習時間': f'{enroll_year}2 / 大一下',
                    }
                },
                '我': {
                    '文學經典閱讀': {
                        '期程': '半',
                        '學分數': 2,
                        '修習時間': f'{enroll_year}1 / 大一上',
                    },
                    '語文與修辭': {
                        '期程': '半',
                        '學分數': 2,
                        '修習時間': f'{enroll_year}2 / 大一下',
                    }
                }
            }
        elif key == '學系必修':
            temp_dict = {
                '微積分(上)': {
                    '期程': '半',
                    '學分數': 3,
                    '修習時間': f'{enroll_year}1 / 大一上',
                    '擋修科目': None,
                    '續修條件': None
                },
                '微積分(下)': {
                    '期程': '半',
                    '學分數': 3,
                    '修習時間': f'{enroll_year}2 / 大一下',
                    '擋修科目': '微積分(上)',
                    '續修條件': '曾修'
                },
                '普通物理(一)': {
                    '期程': '半',
                    '學分數': 3,
                    '修習時間': f'{enroll_year}1 / 大一上',
                    '擋修科目': None,
                    '續修條件': None
                },
                '普通物理(二)': {
                    '期程': '半',
                    '學分數': 3,
                    '修習時間': f'{enroll_year}2 / 大一下',
                    '擋修科目': '普通物理(一)',
                    '續修條件': '曾修'
                },
                '普通物理實驗(一)': {
                    '期程': '半',
                    '學分數': 1,
                    '修習時間': f'{enroll_year}1 / 大一上', # may be postponed
                    '擋修科目': None,
                    '續修條件': None
                },
                '普通物理實驗(二)': {
                    '期程': '半',
                    '學分數': 1,
                    '修習時間': f'{enroll_year}2 / 大一下', # may be postponed
                    '擋修科目': None,
                    '續修條件': None
                },
                '計算機概論(一)': {
                    '期程': '半',
                    '學分數': 3,
                    '修習時間': f'{enroll_year}1 / 大一上',
                    '擋修科目': None,
                    '續修條件': None
                },
                '計算機概論(二)': {
                    '期程': '半',
                    '學分數': 3,
                    '修習時間': f'{enroll_year}2 / 大一下',
                    '擋修科目': '計算機概論(一)',
                    '續修條件': '曾修'
                },
                '線性代數': {
                    '期程': '半',
                    '學分數': 3,
                    '修習時間': f'{enroll_year}1 / 大一上',
                    '擋修科目': None,
                    '續修條件': None
                },
                '機率與統計': {
                    '期程': '半',
                    '學分數': 3,
                    '修習時間': f'{int(enroll_year)+1}2 / 大二上',
                    '擋修科目': None,
                    '續修條件': None
                },
                '電子學(一)': {
                    '期程': '半',
                    '學分數': 3,
                    '修習時間': f'{int(enroll_year)+1}1 / 大二上',
                    '擋修科目': None,
                    '續修條件': None
                },
                '電子實驗(一)': {
                    '期程': '半',
                    '學分數': 1,
                    '修習時間': f'{int(enroll_year)+1}2 / 大二下', # not sure
                    '擋修科目': None,
                    '續修條件': None
                },
                '電路學(一)': {
                    '期程': '半',
                    '學分數': 3,
                    '修習時間': f'{int(enroll_year)+1}1 / 大二上',
                    '擋修科目': None,
                    '續修條件': None
                },
                '電路實驗(一):': {
                    '期程': '半',
                    '學分數': 1,
                    '修習時間': f'{int(enroll_year)+1}1 / 大二上', # not sure
                    '擋修科目': None,
                    '續修條件': None
                },
                '專題製作(一)': {
                    '期程': '半',
                    '學分數': 1,
                    '修習時間': f'{int(enroll_year)+2}2 / 大三下',
                    '擋修科目': None,
                    '續修條件': None
                },
                '專題製作(二)': {
                    '期程': '半',
                    '學分數': 1,
                    '修習時間': f'{int(enroll_year)+3}1 / 大四上',
                    '擋修科目': None,
                    '續修條件': None
                }
            }
        elif key == '學系選修':
            programs = {
                '生產管理學程': {
                    '所屬學系': '工業與系統工程學系',
                    '對應xlsx名': '工業'
                },
                '品質管理學程': {
                    '所屬學系': '工業與系統工程學系',
                    '對應xlsx名': '工業'
                },
                '經營管理學程': {
                    '所屬學系': '工業與系統工程學系',
                    '對應xlsx名': '工業'
                },
                '半導體學程': {
                    '所屬學系': '電子工程學系',
                    '對應xlsx名': '電子'
                },
                '電路設計學程': {
                    '所屬學系': '電子工程學系',
                    '對應xlsx名': '電子'
                },
                '電力學程': {
                    '所屬學系': '電機工程學系',
                    '對應xlsx名': '電機'
                },
                '控制學程': {
                    '所屬學系': '電機工程學系',
                    '對應xlsx名': '電機'
                },
                '通訊學程': {
                    '所屬學系': ['電子工程學系', '電機工程學系'],
                    '對應xlsx名': '通訊'
                },
                '資訊硬體學程': {
                    '所屬學系': '資訊工程學系',
                    '對應xlsx名': '資工'
                },
                '資訊軟體學程': {
                    '所屬學系': '資訊工程學系',
                    '對應xlsx名': '資工'
                },
                '資訊應用學程': {
                    '所屬學系': '資訊工程學系',
                    '對應xlsx名': '資工'
                }
            }
            print(' 本系共有以下學程:')
            mapping = {}
            for idx, program in enumerate(programs):
                print(f'  {idx+1}. {program} - 所屬學系: {programs[program]["所屬學系"]}')
                mapping[str(idx+1)] = program
            cache = []
            for program in ['主修學程一', '主修學程二', '副修學程']:
                while True:
                    try:
                        p = mapping[input(f'> 請輸入{program}的編號 (如\"1\"表示生產管理學程): ')]
                        if p in cache:
                            raise DuplicateProgramError('  學程重複，請重新輸入。')
                        else:
                            temp_dict[program] = p
                            temp_dict[f'{program}所屬學系'] = programs[p]['所屬學系']
                            cache.append(p)
                        break
                    except DuplicateProgramError as e:
                        print(e)
                    except KeyError:
                        print('  無此編號，請重新輸入。')
            temp_dict['學程細項'] = programs
        elif key == '通識延伸選修':
            temp_dict = {
                '天': { '最低應修學分數': 2 },
                '人': {
                    '最低應修學分數': 2,
                    '電資學院指定必修課程': '工程倫理'
                },
                '物': { '最低應修學分數': 2 },
                '我': { '最低應修學分數': 2 }
            }
        elif key == '自由選修': 
            temp_dict = {
                '最多認列學分數': {
                    '輔系': None,
                    '雙主修': None,
                    '跨領域學分學程': None,
                    '就業學程': None,
                    '微型學程（他系）': None,
                    'PBL課程': None,
                    '磨課師(MOCCs)微學分學程': 6,
                    '專業自主學習學分': 2
                }
            }
        json_dict[key] = temp_dict
    
    # set english ability
    json_dict['英文畢業門檻']: {
        '測驗': {
            '全民英檢 GEPT': '中級初試',
            '多益測驗 TOEIC': 550,
            '托福 TOEFL - 紙筆型態 ITP': 450,
            '托福 TOEFL - 網路型態 IBT': 47,
            '雅思測驗 IELTS': 4.0,
            '劍橋大學英語能力認證分級測驗 (Cambridge Main Suite)': 'Preliminary English Test (PET)',
            '劍橋大學國際商務英語能力測驗 (BULATS)': 'ALTE Level 2'
        },
        '全英語專業課程': {
            '最低修習課程數': 2,
            '不認列的課程': ['英文(一)', '英文(二)', '英語聽講(一)', '英語聽講(二)', '實用英文(一)',\
                '實用英文(二)', '商學院商業英語會話(一)', '商學院商業英語會話(二)', '英語檢定技巧']
        }
    }

    # program rules
    json_dict['學程規定']: {
        '學程性質': {
            '主修學程必修': {
                '科目重複或不足之替代學程性質': ['其它主修學程核心', '副修學程不同必修', '副修學程不同核心'],
                '科目重複或不足替代後畢業學分仍不足之替代學程性質': ['電資學院各系專業課程']
            },
            '主修學程核心': {
                '科目重複或不足之替代學程性質': ['其它主修學程核心', '副修學程不同必修', '副修學程不同核心'],
                '科目重複或不足替代後畢業學分仍不足之替代學程性質': ['電資學院各系專業課程'],
                '科目多修抵免學分認列學程性質': ['同一學程選修']
            },
            '主修學程選修': {
                '科目重複之替代學程性質': ['副修學程不同必修', '副修學程不同選修', '副修學程核心']
            },
            '副修學程必修': {},
            '副修學程核心': {
                '科目多修抵免學分認列學程性質': ['同一學程選修']
            },
            '副修學程選修': {}
        }
    }
    json_dict['電資學院跨領域學分學程最低得修習之本院開設專業課程學分數']: 3
    json_dict['國外或香港澳門地區']: {
        '畢業年級相當國內高中二年級之同級同類學校畢業生': {
            '規定之修業年限內最少應增加之應修畢業學分數': 12
        },
        '進入大學前學歷相當國內高中二年級之同級同校學生': {
            '應修課程之學分（不列入大學畢業學分）': {
                '通識課程學分數': 6,
                '基礎理工數學': 6
            }
        }
    }

    with open(f'{path}/{enroll_year}_基本畢業條件.json', 'w', encoding = 'utf-8') as f:
        f.write(json.dumps(json_dict, indent = 4, ensure_ascii = False))

def chinese_number_to_arabic_number(chinese_number):
    single = { '零': 0, '一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '七': 7, '八': 8, '九': 9 }
    units = {'十': 10, '百': 100}
    
    arabic_number = 0
    current_unit_value = 1
    current_value = 0

    for char in reversed(chinese_number):
        if char in single:
            current_value = single[char]
            arabic_number += current_value * current_unit_value
        elif char in units:
            current_unit_value = units[char]
            if current_value == 0:
                arabic_number += current_unit_value
            current_value = 0
        else:
            raise ValueError(f"Invalid character in Chinese number: {char}")

    return arabic_number

def parse_context(department, program_dict):
    if department == '資工':
        # filter untill keys are only 四大類
        del program_dict['選修']['課程']['資訊系必選四大類']
        del program_dict['選修']['課程']['※課程列表請見表一']
        # and set the details in get_program_info() after "for department_name in department_names:" loop
    elif department == '工業':
        # reversly count
        reversed_keys = list(reversed(list(program_dict['選修']['課程'].keys())))
        if '學程選修學分' in reversed_keys[0]: # 會有"＊此四門課程，至多可認列一門為學程選修學分"，就要倒數
            # split substr between '此' and the first '門', ex. "＊此四門課程，至多可認列一門為學程選修學分" -> '四'
            chinese_number = reversed_keys[0].split('門')[0].split('此')[1] # ex. chinese_num = '一百五十六'
            # convert chinese number to arabic number (ex. '四' -> '4')
            arabic_number = chinese_number_to_arabic_number(chinese_number)
            # add review comment (i.e. reversed_keys[0]) into those {arabic_number} courses
            for i in range(1, arabic_number+1):
                program_dict['選修']['課程'][reversed_keys[i]]['審查備註'] = reversed_keys[0]
            # delete reversed_keys[0] in program_dict['選修']['課程']
            del program_dict['選修']['課程'][reversed_keys[0]]
    
    # parse the course name and the credit
    for cur_type in program_dict.keys():
        if cur_type in ['必修', '核心', '選修']:
            legal_courses_format = {}
            #print(program_dict[cur_type]['課程'])
            prev_course = None
            prev_info = None
            for course, info in program_dict[cur_type]['課程'].items():
                # remove illegal characters in the first and the last char
                formatted_course = re.sub(r'^[^\u4e00-\u9fffA-Za-z0-9]+', '', course) # judge first char is legal
                formatted_course = re.sub(r'[^\u4e00-\u9fffA-Za-z0-9\)）]+$', '', formatted_course) # judge last char is legal
                # if no 資工四大類, set credits and review comments
                if not(cur_type == '選修' and department == '資工'):
                    # set credit
                    match_arabic_number = re.search(r'(\d+)$', formatted_course)
                    if match_arabic_number:
                        info['學分數'] = match_arabic_number.group(1) # set credit
                        formatted_course = formatted_course[:match_arabic_number.start()].strip() # remove arabic number in course name
                    else:
                        info['學分數'] = ''
                    # set review comment
                    if department == '通訊':
                        # deal with formatted_course.endswith('(電子系)') or formatted_course.endswith('(電機系)')
                        match_department = re.search(r'\((電子系|電機系)\)$', formatted_course)
                        if match_department:
                            # set department of course
                            match_part = match_department.group(0) # ex. '(電子系)'
                            department_of_course = match_part.replace('(', '').replace(')', '').rstrip() # ex. '電子系'
                            formatted_course = formatted_course[:-len(match_part)].rstrip() # ex. '通訊系統3(電子系)' -> '通訊系統3'
                            info['審查備註'] = f'開課學系：{department_of_course}'
                            # set credit
                            match_arabic_number = re.search(r'(\d+)$', formatted_course)
                            if match_arabic_number:
                                info['學分數'] = match_arabic_number.group(1) # set credit
                                formatted_course = formatted_course[:match_arabic_number.start()].strip() # remove arabic number in course name
                            else: # ex. original formatted_course = '通訊系統(電子系)'
                                info['學分數'] = ''
                        # deal with '或'
                        if formatted_course.startswith('或') and prev_course:
                            # -----------------------WARNING: 只能處理有一個'或'的情況, 即一次有兩門可以互相替換, 如果超過三門可以替換就會出錯-----------------------
                            # delete '或' in formatted_course
                            formatted_course = formatted_course.replace('或', '')
                            # set review comment
                            info['審查備註'] = info['審查備註'] + f'；可替換{prev_course}' if info['審查備註'] else f'可替換{prev_course}'
                            prev_info['審查備註'] = prev_info['審查備註'] + f'；可替換{formatted_course}' if prev_info['審查備註'] \
                                else f'可替換{formatted_course}'
                            # update previous course in legal_courses_format
                            legal_courses_format[prev_course] = prev_info
                            # reset prev_course and prev_info
                            prev_course = None
                            prev_info = None
                # store the current course info
                legal_courses_format[formatted_course] = info
                # update previous course and previous info
                prev_course = formatted_course
                prev_info = info
            program_dict[cur_type]['課程'] = legal_courses_format
    return program_dict

def parse_df_to_dict(df, department):
    # get indices
    ## 各學程名稱所在的column index
    first_row = df.iloc[0, :].values.tolist()
    first_column = df.iloc[:, 0].values.tolist()
    program_indices = { # column indices
        program: idx for idx, program in enumerate(first_row) if program is not None and '學程' in program
    }

    ## 各學程對應的 必修/核心/選修/備註 欄位的row index
    """
    要記錄每個學程對應的必修/核心/選修/備註的row index
    但判斷是系辦寫備註的格式沒有規律
    必修/核心/選修下方的欄位(cell)可能會著名他的學分數也可能沒有
    因此設定max_str_len = 5
    因為備註欄位的長度目前看來必定>5
    而必修/核心/選修和對應的學分數欄位之長度必定<5
    例如電機-控制學程選修的學分數為\"10學分\"長度為4
    若長度為5則表示所需學分數達三位數 不可能
    但若備註欄位長度<=5就會出錯
    """
    max_str_len = 5
    type_indices = {} # row indices
    for program, program_idx in program_indices.items():
        type_indices[program] = {}
        for column_idx, cell in df.iloc[:, program_idx-1].items():
            if (cell is not None) and \
                ((len(str(cell)) < max_str_len and any(p_type in str(cell) for p_type in ['必修', '核心', '選修'])) \
                    or (len(str(cell)) >= max_str_len)):
                type_indices[program][cell] = int(column_idx) - 1

    # get the credits of each type if is not None else ''
    type_credits = {}
    for program, type_dict in type_indices.items():
        type_credits[program] = {}
        for type_name, type_idx in type_dict.items():
            credit_idx = 0 # just for list comprehension's init
            credit_idx = type_idx + 1 if type_idx + 1 < len(first_column) else type_idx
            credit = df.iloc[credit_idx, program_indices[program] - 1]
            if credit is not None:
                type_credits[program][type_name] = credit
            else:
                type_credits[program][type_name] = ''
    
    # get contents by indices
    department_dict = {}
    for program, type_dict in type_indices.items():
        program_dict = {}
        # iterate 必修/核心/選修 by its index
        for cur_type in type_dict.keys():
            if cur_type in ['必修', '核心', '選修']:
                # get the range of the type
                row_start_idx = type_dict[cur_type]
                row_end_idx = list(type_dict.values())[list(type_dict.keys()).index(cur_type) + 1] \
                    if cur_type != list(type_dict.keys())[-1] else len(df.iloc[:, program_indices[program] - 1])
                course_contents = df.iloc[row_start_idx:row_end_idx, program_indices[program]]
                
                # get (審查)備註 of each course in course_contents by program_indices[program]
                column_start_idx = program_indices[program] + 1
                column_end_idx = [program_indices[next_program] - 1 if program != list(program_indices.keys())[-1] else len(df) - 1 \
                    for next_program in list(program_indices.keys())[list(program_indices.keys()).index(program) + 1:]][0] \
                        if program != list(program_indices.keys())[-1] else len(first_row)
                review_comment_contents = df.iloc[row_start_idx:row_end_idx, column_start_idx:column_end_idx]
                if len(course_contents) != len(review_comment_contents):
                    print(f'>  錯誤：{program} {cur_type} 的課程數量與對應的(審核)備註數量不一致！')
                    return None
                else:
                    # get context(str) of each cell into dictionary
                    course_dict = {}
                    for idx, course in enumerate(course_contents.items()):
                        # merge all cells in review_comment_contents into one string
                        # Replace None with '' and merge all cells into one string
                        strings = [value if value is not None else '' \
                            for value in review_comment_contents.iloc[idx, :].values]
                        if cur_type not in program_dict:
                            if department == '資工':
                                program_dict[cur_type] = {
                                    '課程': {}
                                }
                            else:
                                program_dict[cur_type] = {
                                    '最低應修學分數': type_credits[program][cur_type],
                                    '課程': {}
                                }
                        if course[1] is not None:
                            program_dict[cur_type]['課程'][course[1]] = {
                                '學分數': None,
                                '審查備註': '；'.join(string for string in strings if string != '')
                            }
            else:
                # set review comment
                row_idx = type_dict[cur_type]
                column_idx = program_indices[program] - 1
                program_dict['審查備註'] = df.iloc[row_idx, column_idx]
        department_dict[program] = parse_context(department, program_dict)
    return department_dict

def extract_number_between_keywords(text, start_keyword, end_keyword):
    pattern = rf'{start_keyword}.*?(\d+).*?{end_keyword}'
    match = re.search(pattern, text)
    if match:
        return match.group(1)
    return None

# et the graduate credit of single major and double major in CS
def parse_rules_of_single_and_double_cs(cs_dict):
    credit_mapping = {
        '單資工': {
            '必修': {
                '資訊硬體學程': 0,
                '資訊軟體學程': 0,
                '資訊應用學程': 0
            },
            '核心': None
        },
        '雙資工': {
            '必修+核心': {
                '資訊硬體學程+資訊軟體學程': {
                    'set': set(),
                    'credit': 0
                },
                '資訊硬體學程+資訊應用學程': {
                    'set': set(),
                    'credit': 0
                },
                '資訊軟體學程+資訊應用學程': {
                    'set': set(),
                    'credit': 0
                },
            },
            '選修各學程': None,
            '選修總共': None
        }
    }

    # set graduate credit
    for program, program_dict in cs_dict.items():
        for cur_type, contents in program_dict.items():
            if cur_type == '審查備註' and contents is not None:
                # set graduate credit
                strings = contents.split('\n')
                for string in strings:
                    if '擇一修習者' in string:
                        credit_mapping['單資工']['核心'] = extract_number_between_keywords(string, '核心課程中修畢至少', '學分')
                    elif '擇兩個以上修習者' in string:
                        credit_mapping['雙資工']['選修各學程'] = extract_number_between_keywords(string, '至少各', '學分')
                        credit_mapping['雙資工']['選修總共'] = extract_number_between_keywords(string, '須達', '學分以上')
            elif cur_type == '必修' or cur_type == '核心':
                if cur_type == '必修':
                    for course, course_dict in contents['課程'].items():
                        # set single major: 必修都要全修, 核心不用全修, 選修不用修
                        credit_mapping['單資工'][cur_type][program] += int(course_dict['學分數'])
                # set double major: 必修+核心都要全修, 選修不用全修
                for double_major, double_major_dict in credit_mapping['雙資工']['必修+核心'].items():
                    if program in double_major:
                        for course, course_dict in contents['課程'].items():
                            if course not in double_major_dict['set']:
                                double_major_dict['set'].add(course)
                                double_major_dict['credit'] += int(course_dict['學分數'])
    
    # change type of credit from int into str
    for major, contents in credit_mapping.items():
        for cur_type, program_dict in contents.items():
            if major == '單資工' and cur_type == '必修':
                for program, credit in program_dict.items():
                    credit_mapping[major][cur_type][program] = str(credit)
            elif major == '雙資工' and cur_type == '必修+核心':
                for double_major, double_major_dict in contents[cur_type].items():
                    credit_mapping[major][cur_type][double_major] = str(double_major_dict['credit'])
    
    cs_dict['最低應修學分數'] = credit_mapping
    return cs_dict

def parse_cs_four_types_df_to_dict(df, total_dict):
    # read '四大類'
    start_row_idx = 1
    start_column_idx = 0
    end_row_idx, end_column_idx = df.shape
    while df.iloc[end_row_idx - 1, :].isnull().all():
        end_row_idx -= 1
    while df.iloc[:, end_column_idx - 1].isnull().all():
        end_column_idx -= 1
    
    # write into total_dict
    four_classes = { '審查備註': None }
    for column_idx in range(start_column_idx, end_column_idx):
        type_column_idx = None
        for row_idx in range(start_row_idx, end_row_idx):
            if df.iloc[row_idx, column_idx]:
                cell = str(df.iloc[row_idx, column_idx])
                en_code = re.match(r'^[A-Za-z]+$', cell)
                if '、' in cell:
                    type_column_idx = column_idx
                    type_name = cell.split('、')[1].strip()
                    four_classes[type_name] = {
                        '四大類代碼': None,
                        '課程': []
                    }
                elif en_code:
                    four_classes[type_name]['四大類代碼'] = cell
                elif '※' in cell:
                    # delete illegal characters in the beginning and end in the string
                    pattern = r'^[^\u4e00-\u9fffA-Za-z]+|[^\u4e00-\u9fffA-Za-z]+$'
                    cell = re.sub(pattern, '', cell)
                    if not four_classes['審查備註']:
                        four_classes['審查備註'] = '※'
                    if four_classes['審查備註'] == '※':
                        four_classes['審查備註'] += cell
                    else:
                        four_classes['審查備註'] += f'；{cell}'
                elif type_column_idx == column_idx:
                    # remove illegal characters in the beginning and end
                    cell = re.sub(r'^[^\u4e00-\u9fff.A-Za-z0-9]+', '', cell)
                    cell = re.sub(r'[^\u4e00-\u9fffA-Za-z0-9\)）]+$', '', cell)
                    if cell != '課程名稱':
                        # fix mis-spelled and uppercase the first letter
                        match = re.match(r'^malab', cell)
                        if match:
                            cell = cell.replace(match.group(0), 'Matlab')
                        match = re.match(r'^iOS', cell)
                        if match:
                            cell = cell.replace(match.group(0), 'IOS')
                        four_classes[type_name]['課程'].append(cell)
                # else will be ignored (may be unformatted review comments)
    total_dict['資工']['四大類'] = four_classes

    # set program's four types of CS major
    for program in ['資訊硬體學程', '資訊軟體學程', '資訊應用學程']:
        type_names = list(total_dict['資工'][program]['選修']['課程'].keys())
        total_dict['資工'][program]['選修']['認列四大類'] = {}
        for type_name in type_names:
            match = re.match(r'(.*)\(([^()]*)\)$', type_name)
            if match:
                total_dict['資工'][program]['選修']['認列四大類'][match.group(1)] = match.group(2)
        del total_dict['資工'][program]['選修']['課程']
    return total_dict

# read xlsx files from './Program' and generate corresponding 必修/核心/選修 into json file
def get_program_info(path, enroll_year):
    # load json
    json_dict = json.load(open(f'{path}/{enroll_year}_基本畢業條件.json', 'r', encoding = 'utf-8'))
    # read xlsx
    file_names = os.listdir('./Program')
    department_names = ['資工', '工業', '通訊', '電子', '電機']
    try:
        workbooks = {
            department_name: openpyxl.load_workbook(f'./Program/{file_name}') \
            for file_name in file_names \
                for department_name in department_names \
                    if department_name in file_name
        }
    except PermissionError as e:
        print('> 錯誤：請先關閉excel檔案！')
        return
    if len(list(workbooks.keys())) != 5:
        print('> 錯誤：缺少五大學程規劃表！')
        return
    
    # read xlsx into df and parse into dict
    total_dict = {}
    for department_name in department_names:
        if department_name != '資工':
            ws = workbooks[department_name].active
            df = pd.DataFrame(ws.values, columns = [str(i) for i in range(1, ws.max_column + 1)], index = [str(i) for i in range(1, ws.max_row + 1)])
            total_dict[department_name] = parse_df_to_dict(df, department_name)
        else:
            wb = workbooks[department_name]
            sheetnames = wb.sheetnames
            sorted_sheetnames = sorted(sheetnames, key = lambda name: '四大類' in name) # sort sheetname that has '四大類' to the last
            for sheetname in sheetnames:
                ws = wb[sheetname]
                df = pd.DataFrame(ws.values, columns = [str(i) for i in range(1, ws.max_column + 1)], index = [str(i) for i in range(1, ws.max_row + 1)])
                if '四大類' not in sheetname:
                    total_dict[department_name] = parse_df_to_dict(df, department_name)
                    # parse CS single major and double major
                    total_dict[department_name] = parse_rules_of_single_and_double_cs(total_dict[department_name])
                else:
                    # parse 四大類
                    total_dict = parse_cs_four_types_df_to_dict(df, total_dict)
    
    with open(f'{path}/各學程之必修_核心_選修總表.json', 'w', encoding = 'utf-8') as f:
        f.write(json.dumps(total_dict, indent = 4, ensure_ascii = False))

def generate_basic_course_table(enroll_year):
    path = './Generated'
    os.makedirs(path, exist_ok = True)
    # manual setting
    print('> 正在產生修課規定:')
    if os.path.exists(f'{path}/{enroll_year}_基本畢業條件.json'):
        if input(f'> {path}/{enroll_year}_基本畢業條件.json已存在，是否取代(Y/N)? ') != 'Y':
            return
    
    generate_table(path, enroll_year)
    get_program_info(path, enroll_year)