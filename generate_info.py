# -*- coding: utf-8 -*-
import os
import json
import time
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from student_info import StudentInfo

def load_file(file_path, file_name):
    if os.path.exists(f'{file_path}/{file_name}'):
        if file_name.endswith('.json'):
            return json.load(open(f'{file_path}/{file_name}', 'r', encoding = 'utf-8'))
        elif file_name.endswith('.html'):
            return open(f'{file_path}/{file_name}', 'r', encoding = 'utf-8').read()
    else:
        print(f'> 錯誤：\"{file_path}/{file_name}\"不存在！')
        return None

def get_char_width(ch): # 中文、全型字元寬度為兩倍
    if '\u1100' <= ch <= '\uFFDC' or '\u4e00' <= ch <= '\u9fff':
        return 2
    return 1

def generate(info):
    # init workbook
    excel_file_path = f'./Generated/總表.xlsx'
    wb = Workbook()
    ws = wb.active

    # 1. worksheet 0: status
    ws.title = '修課狀態表'
    max_cell_width = [0 for _ in range(18)] # 18 columns

    ## set head contents
    head = []
    major_header = [f"主修學程一：{info.majors['主修學程一']['所屬學系']} - {info.majors['主修學程一']['學程名稱']}", \
            f"主修學程二：{info.majors['主修學程二']['所屬學系']} - {info.majors['主修學程二']['學程名稱']}", \
                f"副修學程：{info.majors['副修學程']['所屬學系']} - {info.majors['副修學程']['學程名稱']}"]
    head.append([''] * 9 + major_header + ['', ''] + major_header + ['']) # TODO: 將第13個cell設為資工四大類條件
    
    ## set body contents
    body = []
    for idx, (credit_type, course) in enumerate(info.sorted_historical_courses.items(), start = 1): # list
        ### write header
        row_header = [f'{idx}. {credit_type}', '課程代碼', '學分數', '期程', '修畢學期', \
            '課程性質', '分數', '修課狀態', '天人物我類別', \
                '主修學程一之必修 / 核心 / 選修', '主修學程二之必修 / 核心 / 選修', '副修學程之必修 / 核心 / 選修', \
                    '資工四大類類別', '最終認定所屬主修學程', '主修學程一之課程審查備註', '主修學程二之課程審查備註', '副修學程之課程審查備註', '其它備註']
        body.append(row_header)

        ### write content
        for course_name, course_dict in course: # tuple
            course_property = ' / '.join(['磨課師' if p == 'M' else ('PBL' if p == 'P' else p) for p in course_dict['課程性質'] if p != ''])
            cs_four_type = ' / '.join([t for t in course_dict['資工四大類類別'] if t != ''])
            final_four_type = course_dict['最終認定所屬主修學程'] if '最終認定所屬主修學程' in course_dict else ''
            row_content = [course_name, course_dict['課程代碼'], course_dict['學分數'], course_dict['期程'], course_dict['修畢學期'], \
                course_property, course_dict['分數'], course_dict['修課狀態'], course_dict['天人物我類別'], \
                    course_dict['課程所屬學程性質']['主修學程一'], course_dict['課程所屬學程性質']['主修學程二'], course_dict['課程所屬學程性質']['副修學程'], \
                        cs_four_type, final_four_type , \
                            course_dict['審查備註']['主修學程一'], course_dict['審查備註']['主修學程二'], course_dict['審查備註']['副修學程'], '']
            body.append(row_content)
        
        ### write unfinished courses
        if credit_type in info.unfinished_courses and info.unfinished_courses[credit_type] != []:
            body += info.unfinished_courses[credit_type]
        
        ### write total credits
        body.append(['已修畢+正在修習之合計學分數', '', info.sub_total_credits[credit_type]] + [''] * 15)
    
    ## set cell width
    for part in [head, body]:
        for row in part:
            for i, cell in enumerate(row):
                cell_length = sum(get_char_width(c) for c in str(cell))
                if cell_length > max_cell_width[i]:
                    max_cell_width[i] = cell_length
    for i in range(18):
        ws.column_dimensions[chr(65 + i)].width = max_cell_width[i] + 10 # offset
    
    ## write head and set details
    head_start_row_index = 1
    for row_index, row_data in enumerate(head, start = head_start_row_index):
        for column_index, cell_data in enumerate(row_data, start = 1):
            cell = ws.cell(row = row_index, column = column_index, value = cell_data)
        if row_index == 1:
            for column in range(1, 19):
                ws.cell(row = row_index, column = column).fill = PatternFill(start_color = '4BACC6', end_color = '4BACC6', fill_type = 'solid')
                ws.cell(row = row_index, column = column).font = Font(bold = True)
        elif row_data[7] == '未通過':
            for column in range(1, 19):
                ws.cell(row = row_index, column = column).fill = PatternFill(start_color = 'FFFF00', end_color = 'FFFF00', fill_type = 'solid')

    ## write body and set details
    body_start_row_index = head_start_row_index + len(head)
    for row_index, row_data in enumerate(body, start = body_start_row_index):
        for column_index, cell_data in enumerate(row_data, start = 1):
            cell = ws.cell(row = row_index, column = column_index, value = cell_data)
        if row_data[7] == '修課狀態':
            for column in range(1, 19):
                ws.cell(row = row_index, column = column).fill = PatternFill(start_color = 'B7DEE8', end_color = 'B7DEE8', fill_type = 'solid')
                ws.cell(row = row_index, column = column).font = Font(bold = True)
        elif row_data[7] == '正在修習':
            for column in range(1, 19):
                ws.cell(row = row_index, column = column).fill = PatternFill(start_color = '92D050', end_color = '92D050', fill_type = 'solid')
        elif row_data[7] == '未修習':
            for column in range(1, 19):
                ws.cell(row = row_index, column = column).fill = PatternFill(start_color = 'FFFF00', end_color = 'FFFF00', fill_type = 'solid')
        elif row_data[0] == '已修畢+正在修習之合計學分數':
            for column in range(1, 19):
                ws.cell(row = row_index, column = column).fill = PatternFill(start_color = 'FFBDF7', end_color = 'FFBDF7', fill_type = 'solid')
        # check eng
        if '英' in ws.cell(row_index, 6).value:
            ws.cell(row = row_index, column = 6).fill = PatternFill(start_color = 'FF0000', end_color = 'FF0000', fill_type = 'solid')
            ws.cell(row = row_index, column = 6).font = Font(color = 'FFFF00')
    # 2. worksheet 1: future course table
    if info.chose_list != []:
        info.generate_future_courses()
        ws = wb.create_sheet(title = '預排課表')
        max_cell_width = [0 for _ in range(9)] # 9 columns

        ## set head
        head = ['時間', '週一', '週二', '週三', '週四', '週五', '週六', '週日', '非同步遠距']
        body = []

        ## get body msgs
        temp_msgs = {
            'A (07:10 ~ 08:00)': [''] * 7,
            '1 (08:10 ~ 09:00)': [''] * 7,
            '2 (09:10 ~ 10:00)': [''] * 7,
            '3 (10:10 ~ 11:00)': [''] * 7,
            '4 (11:10 ~ 12:00)': [''] * 7,
            'B (12:10 ~ 13:00)': [''] * 7,
            '5 (13:10 ~ 14:00)': [''] * 7,
            '6 (14:10 ~ 15:00)': [''] * 7,
            '7 (15:10 ~ 16:00)': [''] * 7,
            '8 (16:10 ~ 17:00)': [''] * 7,
            'C (17:05 ~ 17:55)': [''] * 7,
            'D (18:00 ~ 18:50)': [''] * 7,
            'E (19:55 ~ 19:45)': [''] * 7,
            'F (19:50 ~ 20:40)': [''] * 7,
            'G (20:45 ~ 21:35)': [''] * 7
        }
        for day in info.future_courses:
            day_idx = head.index(day)
            if day != '非同步遠距':
                for time_slot in info.future_courses[day]:
                    temp_msgs[time_slot][day_idx-1] = info.future_courses[day][time_slot]
        
        ## set body
        basic_len = len(temp_msgs) # 15
        online_courses_len = len(info.future_courses['非同步遠距'])
        bigger_len = max(basic_len, online_courses_len)
        for row_idx in range(bigger_len):
            if online_courses_len <= basic_len:
                if row_idx < online_courses_len:
                    body.append([list(temp_msgs.keys())[row_idx]] + temp_msgs[list(temp_msgs.keys())[row_idx]] + [info.future_courses['非同步遠距'][row_idx]])
                else:
                    body.append([list(temp_msgs.keys())[row_idx]] + temp_msgs[list(temp_msgs.keys())[row_idx]] + [''])
            else:
                if row_idx < basic_len:
                    body.append([list(temp_msgs.keys())[row_idx]] + temp_msgs[list(temp_msgs.keys())[row_idx]] + [info.future_courses['非同步遠距'][row_idx]])
                else:
                    body.append([''] * (len(head) - 1) + [info.future_courses['非同步遠距'][row_idx]])
        
        ## set cell width
        for part in [head, body]:
            for row in part:
                for i, cell in enumerate(row):
                    cell_length = sum(get_char_width(c) for c in str(cell))
                    if cell_length > max_cell_width[i]:
                        max_cell_width[i] = cell_length
        for i in range(9):
            ws.column_dimensions[chr(65 + i)].width = max_cell_width[i] + 10 # offset
        
        ## write head and details
        head_start_row_index = 1
        for column_index, cell_data in enumerate(head, start = head_start_row_index):
            cell = ws.cell(row = 1, column = column_index, value = cell_data)
            for column in range(1, 10):
                ws.cell(row = 1, column = column).fill = PatternFill(start_color = '4BACC6', end_color = '4BACC6', fill_type = 'solid')
                ws.cell(row = 1, column = column).font = Font(bold = True)
        
        ## write body and set details
        body_start_row_index = head_start_row_index + 1
        for row_index, row_data in enumerate(body, start = body_start_row_index):
            for column_index, cell_data in enumerate(row_data, start = 1):
                cell = ws.cell(row = row_index, column = column_index, value = cell_data)
                if column_index == 1:
                    ws.cell(row = row_index, column = column_index).fill = PatternFill(start_color = 'B7DEE8', end_color = 'B7DEE8', fill_type = 'solid')
                    ws.cell(row = row_index, column = column_index).font = Font(bold = True)
                else:
                    ws.cell(row = row_index, column = column_index).fill = PatternFill(start_color = 'FDE9D9', end_color = 'FDE9D9', fill_type = 'solid')
    
    # 3. worksheet 2: major1's credit detail
    ws = wb.create_sheet(title = '主副修學程規劃表')

    ## set credit msgs
    credits_mapping = {}
    if info.majors['主修學程一']['對應xlsx名'] == '資工' and info.majors['主修學程二']['對應xlsx名'] == '資工':
        ## set msgs
        for key, value in info.credit_details['資工']['最低應修學分數']['雙資工']['必修+核心'].items():
            total_credit = value if info.majors['主修學程一']['學程名稱'] in key and info.majors['主修學程二']['學程名稱'] in key else '0'
        credits_mapping['主修學程一'] = {
            '必修': {
                '基本要求': '需全修',
                '額外要求': f'兩主修學程必修+核心合計應修{total_credit}學分'
            },
            '核心': {
                '基本要求': '需全修',
                '額外要求': f'兩主修學程必修+核心合計應修{total_credit}學分'
            },
            '選修': {
                '基本要求': f"最低應修{info.credit_details['資工']['最低應修學分數']['雙資工']['選修各學程']}學分",
                '額外要求': f"兩主修學程選修最低合計應修{info.credit_details['資工']['最低應修學分數']['雙資工']['選修總共']}學分"
            }
        }
        credits_mapping['主修學程二'] = credits_mapping['主修學程一']
    else:
        for major in ['主修學程一', '主修學程二']:
            if info.majors[major]['對應xlsx名'] == '資工':
                credits_mapping[major] = {
                    '必修': {
                        '基本要求': '應全修',
                        '額外要求': f"共應修{info.credit_details['資工']['最低應修學分數']['單資工']['必修'][info.majors[major]['學程名稱']]}學分"
                    },
                    '核心': {
                        '基本要求': f"最低應修{info.credit_details['資工']['最低應修學分數']['單資工']['核心']}學分",
                        '額外要求': ''
                    },
                    '選修': { '基本要求': '', '額外要求': '' } }
            else:
                credits_mapping[major] = {
                    '必修': {
                        '基本要求': f"最低應修{info.credit_details[info.majors[major]['對應xlsx名']][info.majors[major]['學程名稱']]['必修']['最低應修學分數']}學分",
                        '額外要求': ''
                    },
                    '核心': {
                        '基本要求': f"最低應修{info.credit_details[info.majors[major]['對應xlsx名']][info.majors[major]['學程名稱']]['核心']['最低應修學分數']}學分",
                        '額外要求': ''
                    },
                    '選修': {
                        '基本要求': f"最低應修{info.credit_details[info.majors[major]['對應xlsx名']][info.majors[major]['學程名稱']]['選修']['最低應修學分數']}學分",
                        '額外要求': ''
                    }
                }
    credits_mapping['副修學程'] = {
        '必修': { '基本要求': '', '額外要求': '' },
        '核心': { '基本要求': '', '額外要求': '' },
        '選修': { '基本要求': '', '額外要求': '' } }
    
    ## fetch majors info
    majors_info = {}
    for major in ['主修學程一', '主修學程二', '副修學程']:
        majors_info[major] = {}
        for single_type in ['必修', '核心', '選修']:
            majors_info[major][single_type] = []
            if info.majors[major]['對應xlsx名'] != '資工':
                for course, course_dict in info.credit_details[info.majors[major]['對應xlsx名']][info.majors[major]['學程名稱']][single_type]['課程'].items():
                    majors_info[major][single_type].append([course, course_dict['學分數'], course_dict['審查備註']])
            else:
                if single_type != '選修':
                    for course, course_dict in info.credit_details['資工'][info.majors[major]['學程名稱']][single_type]['課程'].items():
                        majors_info[major][single_type].append([course, course_dict['學分數'], course_dict['審查備註']])
                else:
                    tempstr = ''
                    for four_type_name, four_type_code in info.credit_details['資工'][info.majors[major]['學程名稱']]['選修']['認列四大類'].items():
                        tempstr = ' / '.join([tempstr, f'{four_type_name} ({four_type_code})']) if tempstr != '' else f'{four_type_name} ({four_type_code})'
                    majors_info[major][single_type].append([f'認列：{tempstr}', '', ''])
    
    ## set max length of body type
    column_len_mapping = {}
    for cur_type in ['必修', '核心', '選修']:
        column_len_mapping[cur_type] = 0
        for cur_major in ['主修學程一', '主修學程二', '副修學程']:
            if len(majors_info[cur_major][cur_type]) > column_len_mapping[cur_type]:
                column_len_mapping[cur_type] = len(majors_info[cur_major][cur_type])
    
    ## set body
    body = []
    basic_header = [f"主修學程一：{info.majors['主修學程一']['學程名稱']}", '',  '', f"主修學程二：{info.majors['主修學程二']['學程名稱']}", '', '', f"副修學程：{info.majors['副修學程']['學程名稱']}", '', '']
    body.append(basic_header)
    must_header = ['必修課程', credits_mapping['主修學程一']['必修']['基本要求'], credits_mapping['主修學程一']['必修']['額外要求'], '必修課程', credits_mapping['主修學程二']['必修']['基本要求'], credits_mapping['主修學程二']['必修']['額外要求'], '必修課程', credits_mapping['副修學程']['必修']['基本要求'], credits_mapping['副修學程']['必修']['額外要求']]
    core_header = ['核心課程', credits_mapping['主修學程一']['核心']['基本要求'], credits_mapping['主修學程一']['核心']['額外要求'], '核心課程', credits_mapping['主修學程二']['核心']['基本要求'], credits_mapping['主修學程二']['核心']['額外要求'], '核心課程', credits_mapping['副修學程']['核心']['基本要求'], credits_mapping['副修學程']['核心']['額外要求']]
    elect_header = ['選修課程', credits_mapping['主修學程一']['選修']['基本要求'], credits_mapping['主修學程一']['選修']['額外要求'], '選修課程', credits_mapping['主修學程二']['選修']['基本要求'], credits_mapping['主修學程二']['選修']['額外要求'], '選修課程', credits_mapping['副修學程']['選修']['基本要求'], credits_mapping['副修學程']['選修']['額外要求']]
    basic_sub_header = ['課程名稱', '學分數', '審查備註', '課程名稱', '學分數', '審查備註', '課程名稱', '學分數', '審查備註']
    for cur_type, header in zip(['必修', '核心', '選修'], [must_header, core_header, elect_header]):
        body.append(header)
        body.append(basic_sub_header)
        for idx in range(column_len_mapping[cur_type]):
            row_data = []
            for cur_major in ['主修學程一', '主修學程二', '副修學程']:
                if idx < len(majors_info[cur_major][cur_type]):
                    row_data += majors_info[cur_major][cur_type][idx]
                else:
                    row_data += [''] * 3
            body.append(row_data)

    ## fetch cs elect - four types
    all_four_type = {}
    cs_elect = []
    for major in ['主修學程一', '主修學程二', '副修學程']:
        if info.majors[major]['對應xlsx名'] == '資工':
            for four_type_name, four_type_code in info.credit_details['資工'][info.majors[major]['學程名稱']]['選修']['認列四大類'].items():
                temp = []
                temp.append(f'{four_type_name} {four_type_code}')
                temp += info.credit_details['資工']['四大類'][four_type_name]['課程']
                cs_elect.append(temp)
    
    ## align cs elect
    max_cs_column_len = len(max(cs_elect, key = len))
    for idx in range(max_cs_column_len):
        for sub_list in cs_elect:
            if idx < len(sub_list):
                sub_list += [''] * (max_cs_column_len - len(sub_list))
    
    ## set body cell width
    max_cell_width = [0 for _ in range(9)]
    for row in body:
        for i, cell in enumerate(row):
            cell_length = sum(get_char_width(c) for c in str(cell))
            if cell_length > max_cell_width[i]:
                max_cell_width[i] = cell_length
    for i in range(9):
        ws.column_dimensions[chr(65 + i)].width = max_cell_width[i] + 10 # offset
    
    ## draw body
    for row_index, row_data in enumerate(body, start = 1):
        for column_index, cell_data in enumerate(row_data, start = 1):
            cell = ws.cell(row = row_index, column = column_index, value = cell_data)
        if row_index == 1:
            for column in range(1, 10):
                ws.cell(row = row_index, column = column).fill = PatternFill(start_color = '4BACC6', end_color = '4BACC6', fill_type = 'solid')
                ws.cell(row = row_index, column = column).font = Font(bold = True)
        elif '必修課程' in row_data[0] or '核心課程' in row_data[0] or '選修課程' in row_data[0]:
            for column in range(1, 10):
                ws.cell(row = row_index, column = column).fill = PatternFill(start_color = '92CDDC', end_color = '92CDDC', fill_type = 'solid')
                ws.cell(row = row_index, column = column).font = Font(bold = True)
        elif '課程名稱' in row_data[0]:
            for column in range(1, 10):
                ws.cell(row = row_index, column = column).fill = PatternFill(start_color = 'B7DEE8', end_color = 'B7DEE8', fill_type = 'solid')
                ws.cell(row = row_index, column = column).font = Font(bold = True)

    ## if have cs elect, then add to body
    if cs_elect != []:
        column_offset_idx = 9 + 1 # 1 for empty column to split cs four type and majors

        ## set cs elect cell width
        max_cell_width = [0 for _ in range(len(cs_elect))]
        for column_index, column in enumerate(cs_elect):
            for row_index, cell in enumerate(column):
                cell_length = sum(get_char_width(c) for c in str(cell))
                if cell_length > max_cell_width[column_index]:
                    max_cell_width[column_index] = cell_length
        for i in range(len(cs_elect)):
            ws.column_dimensions[chr(65 + column_offset_idx + i)].width = max_cell_width[i] + 10 # offset
        
        ## draw cs elect
        for column_index, column_data in enumerate(cs_elect, start = 1):
            for row_index, cell_data in enumerate(column_data, start = 1):
                cell = ws.cell(row = row_index, column = column_offset_idx + column_index, value = cell_data)
                if row_index == 1:
                    ws.cell(row = row_index, column = column_offset_idx + column_index).fill = PatternFill(start_color = 'B7DEE8', end_color = 'B7DEE8', fill_type = 'solid')
                    ws.cell(row = row_index, column = column_offset_idx + column_index).font = Font(bold = True)
    
    # save workbook
    wb.save(excel_file_path)

# read json files, and generate excel file of course status
def generate_info(enroll_year):
    basic_user_info = load_file('./CYCU-Myself', '選課系統_基本資料.json')
    historical_courses = load_file('./CYCU-Myself', '歷年修課.json')
    total_overview = load_file('./CYCU-Myself', '選課系統_總覽.json')
    course_properties = load_file('./CYCU-Myself', '歷年修課與狀態表.html')
    basic_rules = load_file('./Generated', f'{enroll_year}_基本畢業條件.json')
    credit_details = load_file('./Generated', '各學程之必修_核心_選修總表.json')
    if all([basic_user_info, historical_courses, total_overview, course_properties, basic_rules, credit_details]):
        # initialize
        info = StudentInfo(enroll_year, basic_rules, credit_details)
        info.read(basic_user_info, historical_courses, total_overview, course_properties)
        info.parse()
        info.sort_historical_courses()
        info.set_unfinished_courses()
        #info.print_sorted_historical_courses()
        info.write_sorted_historical_courses()

        # generate
        #print(info.unfinished_courses)
        generate(info)
        print(f' 請查看Generated資料夾內的\"總表.xlsx\"！')
    else:
        print('error')
        return
#generate_info('110')